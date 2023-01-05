import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetname,rowno,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rowno,columnno).value

def writeData(file,sheetname,rowno,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rowno,columnno).value = data
    workbook.save(file)

def fillGreenColor(file,sheetname,rowno,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redFill = PatternFill(start_color="60b212",
                          end_color="60b212",
                          fill_type="solid")
    sheet.cell(rowno, columnno).fill = redFill
    workbook.save(file)

def fillRedColor(file,sheetname,rowno,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redFill = PatternFill(start_color="ff0000",
                          end_color="ff0000",
                          fill_type="solid")
    sheet.cell(rowno, columnno).fill = redFill
    workbook.save(file)